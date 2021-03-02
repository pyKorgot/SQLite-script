from openpyxl import load_workbook

import sql_save_select

xlsx_name = './media/resume_import.xlsx'


def get_xlsx_data():
    """Обращается к файлу resume_export.xlsx
    и возвращает все строки занося каждую как отдельное лицо в бд"""
    wb = load_workbook(xlsx_name)
    ws = wb.active

    data = []
    row_range = ws[f'A2:H{ws.max_row}']

    for row in row_range:
        user = []
        for val in row:
            user.append(val.value)
        data.append(user)
    return data


def main():
    sql_save_select.save_data_sql(get_xlsx_data())
    print('Finally import')


if __name__ == '__main__':
    main()
