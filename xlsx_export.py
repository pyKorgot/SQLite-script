from openpyxl import Workbook

import sql_save_select

xlsx_name = './media/resume_export.xlsx'


def create_header(ws):
    """Создание шапки таблицы, бесполезно, но красиво"""
    headers = ['id', 'second_name', 'first_name', 'patronymic', 'region',
               'city', 'phone', 'email']
    i = 0
    for val in ws['A1:H1'][0]:
        val.value = headers[i]
        i += 1


def save_xlsx_data(data):
    """Получает данные и заносит их в таблицу resume_export"""
    wb = Workbook()
    ws = wb.active

    create_header(ws)

    len_data = len(data) + 1
    row_range = ws[f'A2:H{len_data}']

    for user in data:
        row = data.index(user)
        i = 0
        for val in row_range[row]:
            val.value = user[i]
            i += 1

    wb.save(xlsx_name)


def main():
    save_xlsx_data(sql_save_select.get_sql_data())
    print('Finally export')


if __name__ == '__main__':
    main()
